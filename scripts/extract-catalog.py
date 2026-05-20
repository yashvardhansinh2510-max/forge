import os
import json
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed — run: pip3 install openpyxl pillow")

UPLOAD_DIR = "/Users/yashvardhansinhjhala/forge"
OUTPUT_DIR = "/Users/yashvardhansinhjhala/forge/apps/web/public/products"
CATALOG_JSON = "/Users/yashvardhansinhjhala/forge/scripts/catalog.json"
SUMMARY_TXT = "/Users/yashvardhansinhjhala/forge/scripts/catalog-summary.txt"

FILE_MAP = {
    "BM.xlsx": {
        "category": "FAUCETS",
        "subcategory": "Basin Mixers",
    },
    "TBM.xlsx": {
        "category": "FAUCETS",
        "subcategory": "Tall Basin Mixers",
    },
    "3hole.xlsx": {
        "category": "FAUCETS",
        "subcategory": "3-Hole Basin Mixers",
    },
    "WBM.xlsx": {
        "category": "FAUCETS",
        "subcategory": "Wall-Mounted & Concealed Basin Mixers",
    },
    "Single_lever.xlsx": {
        "category": "FAUCETS",
        "subcategory": "Bath & Concealed Shower Mixers",
    },
    "Spout.xlsx": {
        "category": "FAUCETS",
        "subcategory": "Bath Spouts",
    },
    "kitchen.xlsx": {
        "category": "KITCHEN",
        "subcategory": "Kitchen Mixers",
    },
    "SHOWERS_HANSGROHE.xlsx": {
        "category": "SHOWERS",
        "subcategory": "Overhead Showers & Shower Panels",
    },
    "handshower.xlsx": {
        "category": "SHOWERS",
        "subcategory": "Hand Showers",
    },
    "rail.xlsx": {
        "category": "SHOWERS",
        "subcategory": "Shower Rails, Slide Bars & Showerpipes",
    },
    "Thermostat.xlsx": {
        "category": "THERMOSTATS",
        "subcategory": "Thermostats & Controls",
    },
    "Holder.xlsx": {
        "category": "ACCESSORIES",
        "subcategory": "Shower Holders & Wall Outlets",
    },
    "HFAV.xlsx": {
        "category": "ACCESSORIES",
        "subcategory": "Fittings, Arms & Valves",
    },
    "Showerhose.xlsx": {
        "category": "ACCESSORIES",
        "subcategory": "Shower Hoses",
    },
    "Ceramic.xlsx": {
        "category": "BASINS",
        "subcategory": "Ceramics, WCs & Wash Bowls",
    },
}

# Filename aliases: underscore key → actual filename with spaces
FILENAME_ALIASES = {
    "Single_lever.xlsx": "Single lever.xlsx",
    "SHOWERS_HANSGROHE.xlsx": "SHOWERS HANSGROHE.xlsx",
}

FINISH_MAP = {
    "000": "Chrome",
    "007": "Chrome",
    "008": "Chrome",
    "187": "Chrome (Basic Set)",
    "180": "Chrome (Basic Set)",
    "140": "Brushed Bronze (BBR)",
    "340": "Brushed Black Chrome (BBC)",
    "670": "Matt Black",
    "677": "Matt Black",
    "700": "Matt White",
    "990": "Polished Gold Optic (PGO)",
    "300": "Polished Red Gold (PRG)",
    "310": "Brushed Red Gold (BRG)",
    "330": "Polished Black Chrome (PBC)",
    "820": "Brushed Nickel (BN)",
    "250": "Brushed Gold Optic (BGO)",
    "800": "Steel Optic",
    "950": "Brushed Black",
    "450": "White",
    "400": "White / Chrome",
    "600": "Black / Chrome",
    "570": "Petrol",
    "560": "Pink",
    "640": "Zebra",
    "210": "Lion",
    "540": "Petrol",
}


def detect_brand(desc: str) -> str:
    if not desc:
        return "HANSGROHE"
    d = desc.strip()
    if d.startswith("AX ") or d.startswith("AXOR "):
        return "AXOR"
    return "HANSGROHE"


def parse_mrp(val) -> float | None:
    if not val:
        return None
    s = str(val).strip().replace(",", "")
    try:
        f = float(s)
        return f if f > 0 else None
    except (ValueError, TypeError):
        return None


def resolve_filepath(fname: str) -> str | None:
    actual_fname = FILENAME_ALIASES.get(fname, fname)
    path = os.path.join(UPLOAD_DIR, actual_fname)
    if os.path.exists(path):
        return path
    # also try underscore key directly
    path2 = os.path.join(UPLOAD_DIR, fname)
    if os.path.exists(path2):
        return path2
    return None


def build_image_row_map(ws):
    image_row_map = {}
    for img in ws._images:
        try:
            row_idx = img.anchor._from.row
            image_row_map[row_idx] = img
        except Exception:
            continue
    return image_row_map


def find_header_row(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "Article" in str(cell.value):
                return cell.row - 1  # 0-based
    return 0


def process_file(fname, meta, products, seen, images_saved):
    fpath = resolve_filepath(fname)
    if fpath is None:
        print(f"WARNING: {fname} not found — skipping")
        return False

    print(f"\nProcessing: {fname}")
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active

    image_row_map = build_image_row_map(ws)
    header_row_0based = find_header_row(ws)

    all_rows = list(ws.iter_rows(values_only=True))

    current_section = "General"
    file_count = 0
    dup_count = 0

    for row_0idx, row in enumerate(all_rows):
        if row_0idx <= header_row_0based:
            continue

        col0 = row[0] if len(row) > 0 else None   # Sr. No.
        col1 = row[1] if len(row) > 1 else None   # Article No.
        col2 = row[2] if len(row) > 2 else None   # Description
        col4 = row[4] if len(row) > 4 else None   # MRP

        # Section header: col0 has value, col1 is None
        if col0 is not None and col1 is None:
            val = str(col0).strip()
            if val:
                current_section = val
            continue

        article = col1
        if article is None or str(article).strip() == "":
            continue

        article_str = str(article).strip()

        # Skip header rows
        if "Article" in article_str:
            continue

        # Skip rows where article is letters-only (malformed)
        if article_str.isalpha():
            continue

        desc = str(col2).strip() if col2 else ""
        mrp_val = col4

        finish_code = article_str[-3:] if len(article_str) >= 3 else article_str
        finish_name = FINISH_MAP.get(finish_code, "Chrome")

        has_image = row_0idx in image_row_map

        if article_str in seen:
            print(f"  DUPLICATE: {article_str} in {fname} (first seen in {seen[article_str]}) — overwriting")
            dup_count += 1
        seen[article_str] = fname

        product = {
            "sku": article_str,
            "name": desc,
            "brand": detect_brand(desc),
            "brandGroup": "HANSGROHE",
            "category": meta["category"],
            "subcategory": meta["subcategory"],
            "section": current_section,
            "mrp": parse_mrp(mrp_val),
            "gstRate": 18,
            "unit": "pcs",
            "finishCode": finish_code,
            "finishName": finish_name,
            "hasImage": has_image,
            "imageFile": f"{article_str}.png" if has_image else None,
            "sourceFile": fname,
        }
        products[article_str] = product
        file_count += 1

        # Save image
        if has_image:
            img = image_row_map[row_0idx]
            try:
                data = img._data()
                if data[:4] == b'\xd7\xcd\xc6\x9a':
                    continue
                if data[:4] == b'\x89PNG' or data[:3] == b'\xff\xd8\xff':
                    filepath = os.path.join(OUTPUT_DIR, f"{article_str}.png")
                    with open(filepath, 'wb') as f:
                        f.write(data)
                    images_saved[0] += 1
            except Exception as e:
                print(f"  IMG ERROR {article_str}: {e}")
                continue

    print(f"  → {file_count} products, {len(image_row_map)} images in sheet, section: last='{current_section}'")
    return True


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(os.path.dirname(CATALOG_JSON), exist_ok=True)

    products = {}   # sku → product dict (last occurrence wins)
    seen = {}       # sku → first fname
    images_saved = [0]
    files_processed = 0
    files_missing = 0
    dup_total = 0

    for fname, meta in FILE_MAP.items():
        ok = process_file(fname, meta, products, seen, images_saved)
        if ok:
            files_processed += 1
        else:
            files_missing += 1

    # Count duplicates: products where seen[sku] != sourceFile
    for sku, prod in products.items():
        if seen[sku] != prod["sourceFile"]:
            dup_total += 1

    product_list = list(products.values())

    # Write catalog.json
    with open(CATALOG_JSON, 'w', encoding='utf-8') as f:
        json.dump(product_list, f, indent=2, ensure_ascii=False)

    # Stats
    total = len(product_list)
    with_mrp = sum(1 for p in product_list if p["mrp"] is not None)
    without_mrp = total - with_mrp
    imgs = images_saved[0]

    by_brand = {}
    by_cat = {}
    by_subcat = {}
    for p in product_list:
        by_brand[p["brand"]] = by_brand.get(p["brand"], 0) + 1
        by_cat[p["category"]] = by_cat.get(p["category"], 0) + 1
        by_subcat[p["subcategory"]] = by_subcat.get(p["subcategory"], 0) + 1

    lines = [
        "╔═══════════════════════════════════════════╗",
        "║    HANSGROHE CATALOG EXTRACTION REPORT    ║",
        "╠═══════════════════════════════════════════╣",
        f"║  Files processed:    {files_processed:<23}║",
        f"║  Files missing:      {files_missing:<23}║",
        f"║  Total products:     {total:<23}║",
        f"║  Duplicates removed: {dup_total:<23}║",
        f"║  With MRP:           {with_mrp:<23}║",
        f"║  Without MRP (N/A):  {without_mrp:<23}║",
        f"║  With images saved:  {imgs:<23}║",
        "╠═══════════════════════════════════════════╣",
        "║  BY BRAND                                 ║",
        f"║    HANSGROHE:        {by_brand.get('HANSGROHE', 0):<23}║",
        f"║    AXOR:             {by_brand.get('AXOR', 0):<23}║",
        "╠═══════════════════════════════════════════╣",
        "║  BY CATEGORY                              ║",
        f"║    FAUCETS:          {by_cat.get('FAUCETS', 0):<23}║",
        f"║    SHOWERS:          {by_cat.get('SHOWERS', 0):<23}║",
        f"║    THERMOSTATS:      {by_cat.get('THERMOSTATS', 0):<23}║",
        f"║    ACCESSORIES:      {by_cat.get('ACCESSORIES', 0):<23}║",
        f"║    KITCHEN:          {by_cat.get('KITCHEN', 0):<23}║",
        f"║    BASINS:           {by_cat.get('BASINS', 0):<23}║",
        "╠═══════════════════════════════════════════╣",
        "║  BY SUBCATEGORY                           ║",
    ]
    for subcat, cnt in sorted(by_subcat.items(), key=lambda x: -x[1]):
        label = f"    {subcat[:32]}"
        lines.append(f"║  {subcat[:38]:<38} {cnt:<4}║")
    lines += [
        "╠═══════════════════════════════════════════╣",
        "║  IMAGES → apps/web/public/products/       ║",
        "╚═══════════════════════════════════════════╝",
    ]

    report = "\n".join(lines)
    print("\n" + report)

    with open(SUMMARY_TXT, 'w') as f:
        f.write(report + "\n")

    print(f"\ncatalog.json  → {CATALOG_JSON}")
    print(f"summary       → {SUMMARY_TXT}")
    print(f"images        → {OUTPUT_DIR}/")


if __name__ == "__main__":
    main()
